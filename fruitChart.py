print("derave1577")

#import required libraries for this guided practice
import os
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

#imports the function chdir from the library os
from os import chdir

#imports the function load_workbook from the library openpyxl
from openpyxl import load_workbook

#imports reference, BarChart from the sub-library openpyxl.chart
from openpyxl.chart import Reference,BarChart

#change directory to where the Excel file is saved
os.chdir("C:\\PythonFiles")

#wb variable of class openpyxl Workbook is created
#wb is a reference to the Excel file example.xlsx
wb = openpyxl.load_workbook("example.xlsx")

'''
example.xlsx contains one worksheet
You can also use the syntax ws = wb["Sheet"] to get access
to this sheet.
'''

#ws variable of class openpyxl Worksheet created
#ws is a reference to the active worksheet
ws = wb.active

# ws = wb['Sheet1']
# The attribute max_row holds the value
# the attribute max_column holds the value
rows = ws.max_row
cols = ws.max_column

'''
Excel uses letters to refer to column names.
We need the column names to reference cells.
'''

# creates a list of the alphas A-Z based on ASCII
alphas = []
for charCode in range(65,91):
    alphas.append(chr(charCode))

# creates lists to store fruit data for pandas and graphing
fruits = []
fruitValue = []

# for each row in the range of rows
    # iterate through each column in the range of columns
        # print the value of the cell in row, columns
    # print a new line before starting to print the next row
for row in range(1,rows+1):
    for column in range(cols):
        print(ws[alphas[column]+str(row)].value,end="\t")
    print("\n")

for row in range(2,rows+1):
    current_value = int(ws['C'+str(row)].value)
    new_value = current_value + 10
    fruits.append(str(ws['B'+str(row)].value))
    fruitValue.append(new_value)
    ws['C'+str(row)] = new_value

'''
    The += assignm,ent operator allows you to easily increment
    the value of the variable.
    Example: Assuming the current value of myVar is 7,
    myVar += 10 is equivalent to myVar = myVar + 10
    and will result in the value of myVar becoming 17.
'''

# saves the workbook as a new Excel file called fruitData.xlsx
wb.save("fruitData.xlsx")

# get the labels
myDataLabels = Reference(ws,min_col=2,min_row=1,max_row=rows)

# get the data
myData = Reference(ws,min_col=3,min_row=1,max_row=rows)

# Label for the acis at the bottom of the chart
xAxisName="Fruit"

# Label for the numerical data
yAxisName = "Quantity"

# create a BarChart object
myChart = BarChart()

# sets the title attribute
myChart.title = "derave1577"

# removes the legend
myChart.legend = None

# adds the data to be plotted
myChart.add_data(myData)

# sets the values that will appear under the bars
myChart.set_categories(myDataLabels)

# sets the horizontal axis name
myChart.x_axis.title = str(xAxisName)

# sets the vertical axis name
myChart.y_axis.title = str(yAxisName)

# adds myChart to Sheet1 with the upper left corner of the chart in "A11"
ws.add_chart(myChart, "A11")

# saves the workbook as a new Excel file
wb.save("fruitChart.xlsx")

# create a data frame for charting the matplotlib
data = {'Fruits': fruits, 'Quantity': fruitValue}
df = pd.DataFrame(data, columns = ['Fruits', 'Quantity'])

# graph Fruit Data and display the chart
df.plot(x = 'Fruits', y = 'Quantity', kind = 'bar')
plt.xlabel("Fruits")
plt.xticks(rotation=0)
plt.title("derave1577")
plt.show()
