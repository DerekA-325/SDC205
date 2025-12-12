import os
from os import chdir
import pandas as pd
import matplotlib.pyplot as plt
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import Reference, PieChart, BarChart
import datetime



def askUser():
    i = 0
    total = 0
    for i in range(5):
        number = int(input("Please enter a number: "))
        total = total + number
    print(f'The sum of your inputs is: {total}')

def askIncome():
    path = "C:\\FinalExam\\final.csv"
    with open(path, 'a') as csvfile:
        for i in range(5):
            name = input("Enter a name: ")
            salary = int(input("Enter their salary: "))
            csvfile.write(f"{name},{salary}\n")

def excelPie():
    
    names = []
    salaries = []
    path="C:\\FinalExam\\final.csv"
    with open(path, 'r') as csvfile:
        plots = csv.reader(csvfile, delimiter = ',')
        # this takes the values stored in the csv file and stores them in an array for each column
        for row in plots:
            names.append(row[0])
            salaries.append(int(row[1]))
    wb = Workbook()
    ws = wb.active
    rows = ws.max_row
    cols = ws.max_column
    ws.title = "Incomes"
    ws.append(["Name", "Income"])
    # appends the csv file's columns with the stored values in the names and salaries arrays
    for name, salary in zip(names, salaries):
        ws.append([name, salary])

    # this creates the pie chart object, and the following lines fill in the data
    myChart = PieChart()
    labels = Reference(ws, min_col = 1, min_row = 2, max_row = len(names) + 1)
    values = Reference(ws, min_col = 2, min_row = 2, max_row = len(salaries) + 1)
    myChart.add_data(values)
    myChart.set_categories(labels)
    date = datetime.date.today()
    myChart.title = f"derave1577 {date}"
    ws.add_chart(myChart, "D4")
    myChart.height = 15
    myChart.width = 15
    wb.save("final.xlsx")

def verticalBar():
    names = []
    salaries = []
    path="C:\\FinalExam\\final.csv"
    with open(path, 'r') as csvfile:
        plots = csv.reader(csvfile, delimiter = ',')
        for row in plots:
            names.append(row[0])
            salaries.append(int(row[1]))

    # this creates the bar chart and fills in the data
    plt.bar(names,salaries,color="g",width=0.72,label="Salary")
    plt.xlabel("Names")
    plt.ylabel("Salaries")

    date = datetime.date.today()
    plt.title = (f"derave1577 {date}")
    plt.legend()
    plt.show()

def main():
    os.chdir("C:\\FinalExam")
    askUser()
    askIncome()
    excelPie()
    verticalBar()
    
main()
