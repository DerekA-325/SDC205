print("derave1577")

import os
from os import chdir
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import Reference, PieChart, BarChart

os.chdir("C:\\PythonFiles")

wb = openpyxl.load_workbook("StatePopulations.xlsx")

ws = wb.active

rows = ws.max_row
cols = ws.max_column

alphas = []
for charCode in range(65,91):
    alphas.append(chr(charCode))

for row in range(1, rows+1):
    for column in range(cols):
        print(ws[alphas[column]+str(row)].value, end = "\t")
    print("\n")

wb.save("StatePopulationChart.xlsx")

myDataLabels = Reference(ws,min_col = 1, min_row = 2, max_row = rows)
myData = Reference(ws,min_col = 2, min_row = 2, max_row = rows)
xAxisName = ws['A1'].value
yAxisName = ws['B1'].value

myPieChart = PieChart()
myBarChart = BarChart()

myPieChart.title = "derave1577"
myBarChart.title = "derave1577"

myPieChart.add_data(myData)
myBarChart.add_data(myData)

myBarChart.legend = None

myPieChart.set_categories(myDataLabels)
myBarChart.set_categories(myDataLabels)

ws.add_chart(myPieChart, "A20")
ws.add_chart(myBarChart, "H20")

myPieChart.height = 15
myPieChart.width = 15
myBarChart.height = 15
myBarChart.width = 15

wb.save("StatePopulationChart.xlsx")

print("Charts have been saved to StatePopulationChart.xlsx. Open the spreadsheet to view the charts.")
